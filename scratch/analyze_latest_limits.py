import openpyxl
import re

def analyze_file_sheets(file_path, is_bw=False):
    print(f"\n===== Analyzing {file_path} =====")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    numeric_sheets = [s for s in wb.sheetnames if s.isdigit()]
    print(f"All numeric sheets: {numeric_sheets}")
    if not numeric_sheets:
        return
    
    # 按照在 Excel 中的順序，最後一個通常是最新的一天
    latest_sheet = numeric_sheets[-1]
    print(f"Latest sheet selected (the last one): {latest_sheet}")
    
    ws = wb[latest_sheet]
    
    # 讀取第四行和第五行，以查看日期
    print(f"Row 4: {[ws.cell(4, i).value for i in range(1, 10)]}")
    print(f"Row 5: {[ws.cell(5, i).value for i in range(1, 10)]}")
    
    if not is_bw:
        # CW:
        for r in range(8, 45):
            item = ws.cell(r, 1).value
            limit = ws.cell(r, 6).value
            if item:
                print(f"Row {r:02d}: {str(item).strip()} -> Limit: {str(limit).strip() if limit is not None else 'None'}")
    else:
        # BW:
        blocks = [
            ("DMP & LP (Row 8-16)", range(8, 16), 'A', 'D', 'G'),
            ("DEA & SS (Row 19-29)", range(19, 29), 'A', 'D', 'G'),
            ("BFW & MS (Row 31-39)", range(31, 39), 'A', 'D', 'G'),
            ("CBD & CD (Row 41-53)", range(41, 53), 'A', 'D', 'G'),
            ("Right DMP & LP (Row 8-16)", range(8, 16), 'I', 'L', 'O'),
            ("Right DEA & SS (Row 19-29)", range(19, 29), 'I', 'L', 'O'),
            ("Right BFW & MS (Row 31-39)", range(31, 39), 'I', 'L', 'O'),
            ("Right CBD & CD (Row 41-53)", range(41, 53), 'I', 'L', 'O'),
        ]
        for name, r_range, item_col, limit_col1, limit_col2 in blocks:
            print(f"\n* {name}:")
            for r in r_range:
                item = ws[f"{item_col}{r}"].value
                lim1 = ws[f"{limit_col1}{r}"].value
                lim2 = ws[f"{limit_col2}{r}"].value
                if item:
                    print(f"  Row {r:02d}: {str(item).strip()} -> Left Limit (DMP/DEA/BFW/CBD): {str(lim1).strip() if lim1 is not None else 'None'} | Right Limit (LP/SS/MS/CD): {str(lim2).strip() if lim2 is not None else 'None'}")

analyze_file_sheets("d:/WTCA/ref/中龍-CW-0623.xlsx", is_bw=False)
analyze_file_sheets("d:/WTCA/ref/中龍BW -0623.xlsx", is_bw=True)
