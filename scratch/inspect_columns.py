import openpyxl

def inspect_file(file_path, sheets_to_inspect):
    print(f"\n===== Inspecting {file_path} =====")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # 取第一個數字 sheet
    sheet_name = None
    for name in wb.sheetnames:
        if name.isdigit():
            sheet_name = name
            break
    if not sheet_name:
        print("No numeric sheet found")
        return
    
    ws = wb[sheet_name]
    print(f"Inspected sheet name: {sheet_name}")
    
    # 印出前 30 行，前 16 欄 (A ~ P)
    for r in range(1, 40):
        row_vals = []
        for c in range(1, 17):
            val = ws.cell(r, c).value
            row_vals.append(str(val) if val is not None else "")
        # 如果整行都是空的就不印
        if any(row_vals):
            print(f"Row {r:02d}: {row_vals}")

inspect_file("d:/WTCA/ref/中龍-CW-0623.xlsx", ["0623"])
inspect_file("d:/WTCA/ref/中龍BW -0623.xlsx", ["0623"])
